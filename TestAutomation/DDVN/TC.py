import json
import requests
import jsonpath
import openpyxl
from DDVN import Library

def test_AddMultipleStudentData():
    try:
        url = "https://thetestingworldapi.com/api/studentsDetails"
        jd = open("D:\\TestAutomation\\GeneralFiles\\RequestsJson.json")
        wb = openpyxl.load_workbook("D:\\TestAutomation\\GeneralFiles\\TestStuData.xlsx")
        json_Request = json.loads(jd.read())
        obj = Library.A("D:\\TestAutomation\\GeneralFiles\\TestStuData.xlsx","StudentDetails")

        cll = obj.fetch_Col_Count()
        roww = obj.fetch_Row_Count()
        keyList = obj.fetch_key_names()




        sh =wb["StudentDetails"]
        row = sh.max_row
        # print(row)
        # jrq = json.loads(jd.read())
        #
        # print(jrq)

        for i in range(2,row+1):

            updateJsonReq = obj.update_request_with_data(i,json_Request,keyList)

            print("-------------------------------------------------------")

            res = requests.post(url,updateJsonReq)

            print(res.text)

            print("--------------------------------------------------------")




            '''cfn = sh.cell(row=i,column=1)
            cmn = sh.cell(row=i, column=2)
            cln = sh.cell(row=i, column=3)
            doy = sh.cell(row=i, column=4)

            jrq ["first_name"] = cfn.value
            jrq["middle_name"] = cmn.value
            jrq["last_name"] = cln.value
            jrq["date_of_birth"] = doy.value

            res = requests.post(url,jrq)
            print(res.text)
            print(res.status_code)
            jpd = jsonpath.jsonpath(res.json(),"id")
            print(jpd)'''

    except Exception as e:
        print(e)