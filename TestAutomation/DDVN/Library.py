import requests
import json
import jsonpath
import openpyxl

class A:

    try:

        def __init__(self,FilePath,Sheet):
            global sh

            wb = openpyxl.load_workbook(FilePath)
            sh = wb[Sheet]

        def fetch_Row_Count(self):


            rows = sh.max_row
            print(rows)
            return rows

        def fetch_Col_Count(self):


            cols = sh.max_column
            print(cols)
            return cols

        def fetch_key_names(self):
            c = sh.max_column
            li = []

            for i in range(1,c+1):

                print(i)

                cell = sh.cell(row=1,column=i)
                li.insert(i-1, cell.value)

            print(li)
            return li

        def update_request_with_data(self,rowNum,jsonRequest,keyList):

            cl = sh.max_column

            for i in range(1,cl+1):

                cell = sh.cell(row=rowNum,column=i)

                jsonRequest[keyList[i-1]] = cell.value

            return jsonRequest

    except Exception as e:
        print(e)