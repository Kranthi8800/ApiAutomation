import json
import requests
import jsonpath
import openpyxl

def test_AddMultipleStudentData():
    try:
        url = "https://thetestingworldapi.com/api/studentsDetails"
        jd = open("D:/TestAutomation/GeneralFiles/RequestsJson.json")
        wb = openpyxl.load_workbook("D:/TestAutomation/GeneralFiles/TestStuData.xlsx")
        sh =wb["StudentDetails"]
        row = sh.max_row
        print(row)
        jrq = json.loads(jd.read())

        print(jrq)

        for i in range(2,row+1):

            cfn = sh.cell(row=i,column=1)
            cmn = sh.cell(row=i, column=2)
            cln = sh.cell(row=i, column=3)
            doy = sh.cell(row=i, column=4)

            jrq ["first_name"] = cfn.value
            jrq["middle_name"] = cmn.value
            jrq["last_name"] = cln.value
            jrq["date_of_birth"] = doy.value

            res = requests.post(url,jrq)
            print(res.text)
            print(res.status_code)

    except Exception as e:
        print(e)